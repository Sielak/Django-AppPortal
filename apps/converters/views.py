from django.shortcuts import render
from django.http import HttpResponseRedirect
from django.http import HttpResponse
from django.core.files.storage import FileSystemStorage
from django.conf import settings
import os
import sqlite3 as sql
from .models import AldiMatrix
import pandas as pd
from django.contrib import messages
import openpyxl
import csv
from sentry_sdk import last_event_id


ALLOWED_EXTENSIONS = set(['db', 'xlsx'])


def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1] in ALLOWED_EXTENSIONS


def db2csv(request):
    upload_folder = settings.MEDIA_ROOT + '/db2csv/'
    if request.method == 'POST':
        try:
            path = os.listdir(upload_folder)
            path2db = upload_folder + path[0]
            os.remove(path2db)
        except IndexError:
            pass
        try:
            file = request.FILES['myfile']
        except KeyError:
            file = ''
        if file and allowed_file(str(file)):
            myfile = request.FILES['myfile']
            fs = FileSystemStorage(location=upload_folder)
            fs.save(myfile.name, myfile)
            con = sql.connect(upload_folder + myfile.name)
            con.row_factory = sql.Row
            cur = con.cursor()
            data = cur.execute("""
            SELECT
                uid,
                session_id,
                event_type,
                event_name,
                strftime('%d-%m-%Y  %H:%M:%S', datetime(timestamp/1000,'unixepoch')) as date
            FROM
                event;
            """)
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="marketing_db.csv"'

            writer = csv.writer(response)
            writer.writerow(['id', 'Session ID', 'Event type',
                             'Event name', 'Timestamp'])
            writer.writerows(data)

            return response

        return HttpResponseRedirect('db2csv')
    else:

        return render(request, 'converters/db2csv.html')


def aldi_orders(request):
    if request.method == 'POST':
        location = request.POST['location']
        size = request.POST['size']
        file_type = request.POST['type']
        column = location + "_" + size
        rows = AldiMatrix.objects.values_list(
            'jeeves_code', 'unit_cost', 'type', column)  # Select only 4 columns
        filter_kwargs = {
            "{}__gt".format(column): 0
        }
        # Filter results. Column > 0
        rows_filtered = rows.filter(**filter_kwargs)
        mts = []
        mto = []
        for item in rows_filtered:
            if 'MTS' in item:
                mts.append(item)
            else:
                mto.append(item)

        if file_type == 'MTS':
            list_type = mts
        else:
            list_type = mto

        # Download Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename={0}.xlsx'.format(
            file_type)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sheet1'

        row_num = 0

        columns = [
            (u"JvsCode", 15),
            (u"Quantity", 70),
            (u"PoD", 70),
            (u"Price", 70),
            (u"ExtTxt", 70),
            (u"OrdSpecTxt", 70),
            (u"TxtBetweenRows", 70)
        ]

        for col_num in range(len(columns)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = columns[col_num][0]

        for item in list_type:
            row_num += 1
            row = [
                item[0],
                item[3],
                '',
                item[1],
            ]
            for col_num in range(len(row)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = row[col_num]

        wb.save(response)
        return response

    else:

        context = {
            'test1': 'Site for Aldi orders',
        }

        return render(request, 'converters/aldi_orders.html', context)


def aldi_orders_admin(request):
    upload_folder = settings.MEDIA_ROOT + '/aldi/'
    if request.method == 'POST':
        try:
            path = os.listdir(upload_folder)
            path2file = os.path.abspath('media/aldi/{0}'.format(path[0]))
            os.remove(path2file)
        except (IndexError, FileNotFoundError) as e:
            print(e)
        try:
            file = request.FILES['myfile']
        except KeyError:
            file = ''
        if file and allowed_file(str(file)):
            myfile = request.FILES['myfile']
            fs = FileSystemStorage(location=upload_folder)
            fs.save('aldi.xlsx', myfile)
            AldiMatrix.objects.all().delete()  # truncate table
            path2file = upload_folder + 'aldi.xlsx'
            df = pd.read_excel(path2file, sheet_name='Main', skiprows=2)
            # test1 = {df.columns.get_loc(c):c for idx, c in enumerate(df.columns)}
            # print(test1)
            df1 = df[['Jeeves Code',
                      'Unit Cost',
                      'Type',
                      'e_850',
                      'e_900',
                      'e_940',
                      'e_960',
                      'e_990_1006',
                      'e_v2_1125_1140',
                      'e_v2_1254_1315',
                      'erh_850',
                      'erh_900',
                      'erh_940',
                      'erh_960',
                      'erh_990_1006',
                      'erh_v2_1125_1140',
                      'erh_v2_1254_1315',
                      's_850',
                      's_900',
                      's_940',
                      's_960',
                      's_990_1006',
                      's_v2_1125_1140',
                      's_v2_1254_1315',
                      'srh_850',
                      'srh_900',
                      'srh_940',
                      'srh_960',
                      'srh_990_1006',
                      'srh_v2_1125_1140',
                      'srh_v2_1254_1315',
                      'w_850',
                      'w_900',
                      'w_940',
                      'w_960',
                      'w_990_1006',
                      'w_v2_1125_1140',
                      'w_v2_1254_1315',
                      'wrh_850',
                      'wrh_900',
                      'wrh_940',
                      'wrh_960',
                      'wrh_990_1006',
                      'wrh_v2_1125_1140',
                      'wrh_v2_1254_1315'
                      ]]
            # Drop rows with NaN in Jeeves code
            df1 = df1.dropna(subset=['Jeeves Code'])
            df1['Jeeves Code'] = df1['Jeeves Code'].astype(int)
            for index, row in df1.iterrows():
                model = AldiMatrix()
                model.jeeves_code = row['Jeeves Code']
                model.unit_cost = row['Unit Cost']
                model.type = row['Type']
                model.e_850 = row['e_850']
                model.e_900 = row['e_900']
                model.e_940 = row['e_940']
                model.e_960 = row['e_960']
                model.e_990_1006 = row['e_990_1006']
                model.e_v2_1125_1140 = row['e_v2_1125_1140']
                model.e_v2_1254_1315 = row['e_v2_1254_1315']
                model.erh_850 = row['erh_850']
                model.erh_900 = row['erh_900']
                model.erh_940 = row['erh_940']
                model.erh_960 = row['erh_960']
                model.erh_990_1006 = row['erh_990_1006']
                model.erh_v2_1125_1140 = row['erh_v2_1125_1140']
                model.erh_v2_1254_1315 = row['erh_v2_1254_1315']
                model.s_850 = row['s_850']
                model.s_900 = row['s_900']
                model.s_940 = row['s_940']
                model.s_960 = row['s_960']
                model.s_990_1006 = row['s_990_1006']
                model.s_v2_1125_1140 = row['s_v2_1125_1140']
                model.s_v2_1254_1315 = row['s_v2_1254_1315']
                model.srh_850 = row['srh_850']
                model.srh_900 = row['srh_900']
                model.srh_940 = row['srh_940']
                model.srh_960 = row['srh_960']
                model.srh_990_1006 = row['srh_990_1006']
                model.srh_v2_1125_1140 = row['srh_v2_1125_1140']
                model.srh_v2_1254_1315 = row['srh_v2_1254_1315']
                model.w_850 = row['w_850']
                model.w_900 = row['w_900']
                model.w_940 = row['w_940']
                model.w_960 = row['w_960']
                model.w_990_1006 = row['w_990_1006']
                model.w_v2_1125_1140 = row['w_v2_1125_1140']
                model.w_v2_1254_1315 = row['w_v2_1254_1315']
                model.wrh_850 = row['wrh_850']
                model.wrh_900 = row['wrh_900']
                model.wrh_940 = row['wrh_940']
                model.wrh_960 = row['wrh_960']
                model.wrh_990_1006 = row['wrh_990_1006']
                model.wrh_v2_1125_1140 = row['wrh_v2_1125_1140']
                model.wrh_v2_1254_1315 = row['wrh_v2_1254_1315']
                try:
                    model.save()
                except ValueError:
                    print('Error in input file in row with Jeeves code:',
                          row['Jeeves Code'])
                    messages.info(request,
                                  'Error in input file in row with Jeeves code: {0}'.format(row['Jeeves Code']))
                    pass

        return HttpResponseRedirect('aldi_orders_admin')
    else:
        try:
            db_list = os.listdir(upload_folder)[0]
        except (IndexError, FileNotFoundError) as e:
            print(e)
            db_list = ''
        context = {
            'test1': 'Site for Aldi upload Admin',
            'db_list': db_list
        }

        return render(request, 'converters/aldi_orders_admin.html', context)

def handler500(request, *args, **argv):
    try:
        event_id = last_event_id()
    except NameError:
        event_id = 'None'
    return render(request, "converters/500.html", {
        'sentry_event_id': event_id,
    }, status=500)
